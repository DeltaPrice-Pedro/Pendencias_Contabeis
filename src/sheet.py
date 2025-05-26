from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tkinter.filedialog import asksaveasfilename

from PySide6.QtCore import QObject, Signal

class Sheet(QObject):
    """
    Responsável por gerar e salvar relatórios de envios em formato Excel (.xlsx).
    """
    start = Signal()
    end = Signal()
    result = Signal(str)

    def __init__(self, content: dict[str, list]):
        super().__init__()
        self.file_path = ''
        self.default_extension = '.xlsx'
        self.title_upload = 'Defina onde salvar o relatório'
        self.file_types = (
            ('Planilha Excel', '*.xlsx'),
            ('Todos os Arquivos', '*.*')
        )
        self.content = content

        self.text_font = Font(
                name='Calibri',
                size=16,
                bold=True,
                color='FF000000'
            )
        
        self.thin_border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000')
        )

        self.dashed_border = Border(
            left=Side(border_style='dashed', color='FF000000'),
            right=Side(border_style='dashed', color='FF000000'),
            top=Side(border_style='dashed', color='FF000000'),
            bottom=Side(border_style='dashed', color='FF000000')
        )

        self.center_align = Alignment(horizontal='center')
        self.left_align = Alignment(horizontal='left')


    def write(self) -> dict:
        """
        Cria e salva o arquivo Excel com os dados fornecidos.
        """
        self.start.emit()

        wb = Workbook()
        ws = wb.active
        for index_column, values in enumerate(self.content.keys(), 1):
            cell = ws.cell(1, index_column, values)
            cell.font = self.text_font
            cell.border = self.dashed_border
            cell.alignment = self.center_align
            ws.column_dimensions[get_column_letter(index_column)].width = 40

        for index_column, values in enumerate(self.content.values(), 1):
            for index_row, value in enumerate(values, 2):
                cell = ws.cell(index_row, index_column, value)
                cell.border = self.thin_border
                cell.alignment = self.left_align

        wb.save(self.file_path)
        
        self.result.emit(self.file_path)
        self.end.emit()

    def upload(self):
        """
        Abre o diálogo para o usuário escolher onde salvar o arquivo Excel.
        Returns:
            str: Caminho do arquivo salvo.
        """
        self.file_path = asksaveasfilename(
            title= self.title_upload, 
            defaultextension= self.default_extension,
            filetypes= self.file_types
        )

        if self.file_path == '':
            raise Exception('Operação cancelada')

        return self.file_path